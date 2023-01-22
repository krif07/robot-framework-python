import openpyxl

file = openpyxl.load_workbook("C:\\Users\\krif0\\PycharmProjects\\UdemyrobotFramework\\Test_Case\\Python_Excel\\Referencia.xlsx")
print(file)
print(file.sheetnames)
print(f"Active page -> {file.active.title}")

active_page = file.active
print(active_page['A1'].value)
print(active_page['A2'].value)
print(active_page['B1'].value)
print(active_page['B2'].value)
print(active_page['C1'].value)
print(active_page['C2'].value)

page2 = file['2-Producto']
print(page2['A1'].value)
print(page2['A2'].value)
print(page2['B1'].value)

cell = active_page.cell(7, 2)
print(f"cell: {cell.value}")

#
print("*"*50)

rows = active_page.max_row
columns = active_page.max_column

print(f"rows -> {rows} -- columns -> {columns}")
for row in active_page['A1':'C9']:
    for column in row:
        print(column.value)
    print("-"*20)

print("*"*50)
for row in range(1, rows+1):
    for column in range(1, columns+1):
        cell = active_page.cell(row, column)
        print(cell.value)
    print("-" * 20)

print("*"*50)

# Write values in sheet
result = active_page['C3'].value = "OK"
result = active_page['C4'].value = "OK"
result = active_page['C5'].value = "KO"

file.save("C:\\Users\\krif0\\PycharmProjects\\UdemyrobotFramework\\Test_Case\\Python_Excel\\Referencia2.xlsx")